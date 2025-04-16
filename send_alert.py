import pandas as pd
import re
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Thresholds by category
THRESHOLDS = {
    'Accessories': 200,
    'Laptop': 150,
    'Monitor': 30
}

def send_low_inventory_alert(file_path='inventory_data.xlsx', report_path='alert_report.xlsx'):
    df = pd.read_excel(file_path)

    grouped = (
        df.groupby(['Category', 'Brand', 'Model'])['Quantity']
        .sum()
        .reset_index()
    )

    low_inventory = []
    sufficient_inventory = []

    for _, row in grouped.iterrows():
        category = row['Category'].strip()
        brand = row['Brand'].strip()
        model = row['Model'].strip()
        model = re.sub(re.escape(brand), '', model, flags=re.IGNORECASE).strip()
        brand_model = f"{brand} {model}".strip()
        quantity = row['Quantity']
        threshold = THRESHOLDS.get(category, 20)

        if quantity < threshold:
            low_inventory.append({
                'Category': category,
                'Device': brand_model,
                'Quantity': quantity,
                'Threshold': threshold
            })
        else:
            sufficient_inventory.append(f"✅ {category}: {brand_model} – {quantity} units")

    # Create bar chart in memory
    labels, quantities, colors = [], [], []
    for _, row in grouped.iterrows():
        category = row['Category'].strip()
        brand = row['Brand'].strip()
        model = row['Model'].strip()
        model = re.sub(re.escape(brand), '', model, flags=re.IGNORECASE).strip()
        brand_model = f"{brand} {model}".strip()
        labels.append(f"{category}: {brand_model}")
        qty = row['Quantity']
        quantities.append(qty)
        colors.append('red' if qty < THRESHOLDS.get(category, 20) else 'green')

    fig_width = max(10, 0.5 * len(labels))
    fig, ax = plt.subplots(figsize=(fig_width, 6))
    ax.bar(range(len(labels)), quantities, color=colors)
    ax.set_title('Remaining Inventory by Device Model')
    ax.set_xlabel('Device Model')
    ax.set_ylabel('Quantity')
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right')
    fig.tight_layout()

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=150)
    plt.close(fig)
    buffer.seek(0)

    # Write report with chart and sheets
    with pd.ExcelWriter(report_path, engine='openpyxl', mode='w') as writer:
        if low_inventory:
            pd.DataFrame(low_inventory).to_excel(writer, sheet_name='Low Inventory', index=False)

        if sufficient_inventory:
            df_sufficient = pd.DataFrame([{'Status': line} for line in sufficient_inventory])
            df_sufficient.to_excel(writer, sheet_name='Sufficient Inventory', index=False)

        # Create chart sheet
        writer.book.create_sheet(title='Inventory Chart')
        chart_ws = writer.book['Inventory Chart']
        img = Image(buffer)
        chart_ws.add_image(img, 'A1')

        # Remove default sheet if exists
        if 'Sheet' in writer.book.sheetnames:
            del writer.book['Sheet']

    print(f"\n📁 Alert report saved to: {report_path}")

    if low_inventory:
        print("⚠️ Warning: The following devices are running low on stock:")
        for item in low_inventory:
            print(f"- {item['Category']}: {item['Device']} – {item['Quantity']} units (threshold: {item['Threshold']})")
    else:
        print("✅ All devices are sufficiently stocked.")

    print("\n--- Devices sufficiently stocked ---")
    for line in sufficient_inventory:
        print(line)

def send_email(subject, body, to_emails, attachment_path):
    sender_email = 'dongtrieuvohoang@gmail.com'
    app_password = 'fhjq podp xgvy ybdj'  # Replace with your App Password

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(to_emails)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    with open(attachment_path, 'rb') as attachment:
        part = MIMEApplication(attachment.read(), _subtype="xlsx")
        part.add_header('Content-Disposition', 'attachment', filename=attachment_path)
        msg.attach(part)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, app_password)
        server.sendmail(sender_email, to_emails, msg.as_string())
        server.quit()
        print("✅ Email sent successfully.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")

def is_valid_email(email):
    email_pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.(com|net|org|edu|gov|co|info)$'
    return bool(re.match(email_pattern, email))

if __name__ == "__main__":
    send_low_inventory_alert()

    emails_input = input("Enter recipient email addresses (separated by commas): ").strip()
    email_list = [e.strip() for e in emails_input.split(",") if is_valid_email(e)]

    if not email_list:
        print("❌ No valid emails found. Please try again.")
    else:
        subject = "Inventory Alert Report"
        body = "Hi team,\n\nThis is the latest inventory report. Please find the attached Excel file.\n\nBest regards,\nTrieu Vo"
        send_email(subject, body, email_list, 'alert_report.xlsx')
